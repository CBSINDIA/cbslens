import sys
import pandas as pd
import json
from openpyxl import Workbook

wb = Workbook()
ws = wb.active


arg = sys.argv[1]
fname = "D:\\Python\\untitled1\\"
fname += arg + ".json"
with open(fname, 'r') as f:
  distros_dict = json.load(f)

d = []
k = []
a = 1

# print (type(distros_dict))
if isinstance(distros_dict, dict):
  for blocks in distros_dict['Blocks']:
    if blocks['BlockType'] == 'LINE':
      d = [blocks['Page'], blocks['Text'], blocks['Geometry']['BoundingBox']['Top'],
           blocks['Geometry']['BoundingBox']['Left'], blocks['Geometry']['BoundingBox']['Width'],
           blocks['Geometry']['BoundingBox']['Height']]
      k.append(d)
else:
  for lis in distros_dict:
    for blocks in lis['Blocks']:
      if blocks['BlockType'] == 'LINE':
        d = [blocks['Page'], blocks['Text'], blocks['Geometry']['BoundingBox']['Top'],
             blocks['Geometry']['BoundingBox']['Left'], blocks['Geometry']['BoundingBox']['Width'],
             blocks['Geometry']['BoundingBox']['Height']]
        k.append(d)

dk = pd.DataFrame(k)
dk.columns = ['Page','Text','Top','Left','width','height']
dk["line"] = int(0)
lin = 1
for row in dk.index:
  if row > 0:
    if (dk['Page'][row] > dk['Page'][row-1]):
      lin = 1
    #if (dk['Page'][row] == 3):
      #print(dk['Top'][row-1],dk['Top'][row],(dk['Top'][row] - dk['Top'][row - 1]),dk['Text'][row - 1])
    if (dk['Top'][row] - dk['Top'][row-1]) >= 0.010 :#or (dk['Left'][row-1] - dk['Left'][row] > 0.50) :
      lin += 1
    #dk["line"][row] = lin
    dk._set_value(row,["line"],lin)
  else:
    dk._set_value(row, ["line"], lin)
    #dk["line"][row] = lin

dk['col'] = int(0)
col = 0
inc = 0
for row in dk.index:
  if row > 0:
    if (dk["line"][row] > dk["line"][row-1]) or (dk['Page'][row] > dk['Page'][row-1]):
      col = 1
      inc += 1
      #dk['col'][inc] =  col
      dk._set_value(inc, ["col"], col)
    else:
      inc += 1
      col += 1
      #dk['col'][inc] =  col
      dk._set_value(inc, ["col"], col)
  else:
    col = 1
    #dk['col'][inc] = col
    dk._set_value(inc, ["col"], col)

#print(dk.query('line==1 and col==1' )['Text'])
ws = wb.create_sheet("Page1")

for row in dk.index:
  #if dk['Page'][]
  x = int(dk['line'][row])
  y = int(dk['col'][row])

  str1 = dk.query('line==@x and col==@y')['Text'].iloc[0]
  #ws.cell(row=x, column=y).value = dk.query('line==@x and col==@y'['Text'].iloc[0])
  ny = int(dk['Left'][row] * 13)
  ny = round(ny)
  #print(dk['Left'][row],ny,str)
  if ny <=0:
    ny = 1
  else:
    ny +=1

  if row > 0 and dk['Page'][row] > dk['Page'][row-1]:
    ws = wb.create_sheet("Page"+str(dk["Page"][row]))

  ws.cell(row=x, column=ny).value = str1
  #print(dk['Left'][row], ny, str)

wb.save('D:\\Python\\untitled1\\'+arg+'.xlsx')
#cfname = "D:\\Python\\untitled1\\"
#cfname += arg + ".csv"
#dk.to_csv(cfname, sep='|', header=False)


#print(dk.head(10))
#print(dk)
#dk[1] = dk[1].str.upper()
